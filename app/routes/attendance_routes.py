from flask import Blueprint, render_template, request, session, redirect, url_for, flash, send_file
from datetime import datetime, timedelta, date

from app import db
from app.models import Attendance, Student

from io import BytesIO
from openpyxl import Workbook

attendance_bp = Blueprint("attendance", __name__)

@attendance_bp.route("/admin/attendance")
def attendance_list():

    if "admin_id" not in session:
        return redirect(url_for("main.admin_login"))

    # Get selected date from URL
    selected_date = request.args.get("date")

    # Check if admin wants to see all attendance
    show_all = request.args.get("all") == "true"

    # --------------------------------
    # SHOW ALL ATTENDANCE
    # --------------------------------

    if show_all:

        records = Attendance.query.order_by(
            Attendance.time_in.desc()
        ).all()

        display_date = "All Attendance"

        date_value = ""

    # --------------------------------
    # SHOW SELECTED DATE
    # --------------------------------

    else:

        if selected_date:

            try:

                selected_date = datetime.strptime(
                    selected_date,
                    "%Y-%m-%d"
                ).date()

            except ValueError:

                selected_date = datetime.now().date()

        else:

            # Default = TODAY
            selected_date = datetime.now().date()


        records = Attendance.query.filter(
            Attendance.date == selected_date
        ).order_by(
            Attendance.time_in.desc()
        ).all()


        display_date = selected_date.strftime(
            "%d-%m-%Y"
        )

        date_value = selected_date.strftime(
            "%Y-%m-%d"
        )


    return render_template(

        "admin/attendance.html",

        records=records,

        display_date=display_date,

        selected_date=date_value,

        show_all=show_all

    )

@attendance_bp.route("/student/attendance", methods=["GET", "POST"])
def student_attendance():

    if "student_id" not in session:
        return redirect(url_for("main.student_login"))

    student = Student.query.get(session["student_id"])

    if not student:
        session.clear()
        return redirect(url_for("main.student_login"))

    if request.method == "POST":

        today = datetime.now().date()

        # --------------------------------
        # PREVENT MULTIPLE ATTENDANCE
        # --------------------------------

        already_marked = Attendance.query.filter(
            Attendance.student_id == student.student_id,
            Attendance.date == today
        ).first()

        if already_marked:

            flash(
                "You have already marked attendance today.",
                "warning"
            )

            session.clear()

            return redirect(url_for("main.home"))

        # --------------------------------
        # CREATE ATTENDANCE
        # --------------------------------

        attendance = Attendance(

            # Student details from database
            student_id=student.student_id,
            student_name=student.name,

            year=student.year,
            branch=student.branch,
            section=student.section,

            # Topic / system / remarks from student form
            topic=request.form["topic"],

            system_number=request.form["system_number"],

            remarks=request.form["remarks"],

            # Automatically generated
            date=today,

            time_in=datetime.now()

        )

        db.session.add(attendance)

        db.session.commit()

        flash(
            "Attendance Submitted Successfully!",
            "success"
        )

        # Logout student after attendance
        session.clear()

        return redirect(url_for("main.home"))

    return render_template(
        "student/attendance.html",
        student=student
    )


@attendance_bp.route("/admin/report", methods=["GET", "POST"])
def attendance_report():

    if "admin_id" not in session:
        return redirect(url_for("main.admin_login"))

    records = []

    selected_group = ""
    selected_report_type = "daily"
    student_id = ""

    if request.method == "POST":

        student_id = request.form.get("student_id", "").strip()
        selected_report_type = request.form.get("report_type", "daily")
        selected_group = request.form.get("group", "").strip()

        today = datetime.now().date()

        # Start with all attendance records
        query = Attendance.query

        # --------------------------------
        # STUDENT FILTER
        # --------------------------------
        if student_id:
            query = query.filter(
                Attendance.student_id == student_id
            )

        # --------------------------------
        # GROUP / BRANCH FILTER
        # --------------------------------
        if selected_group:
            query = query.filter(
                Attendance.branch == selected_group
            )

        # --------------------------------
        # DATE FILTER
        # --------------------------------
        if selected_report_type == "daily":

            query = query.filter(
                Attendance.date == today
            )

        elif selected_report_type == "weekly":

            start_date = today - timedelta(days=7)

            query = query.filter(
                Attendance.date >= start_date
            )

        elif selected_report_type == "monthly":

            start_date = today.replace(day=1)

            query = query.filter(
                Attendance.date >= start_date
            )

        records = query.order_by(
            Attendance.date.desc(),
            Attendance.time_in.desc()
        ).all()

    return render_template(
        "admin/report.html",
        records=records,
        selected_group=selected_group,
        selected_report_type=selected_report_type,
        student_id=student_id
    )

@attendance_bp.route("/admin/report/export", methods=["POST"])
def export_report():

    if "admin_id" not in session:
        return redirect(url_for("main.admin_login"))

    # --------------------------------
    # GET FILTER VALUES
    # --------------------------------

    student_id = request.form.get("student_id", "").strip()

    report_type = request.form.get(
        "report_type",
        "daily"
    )

    selected_group = request.form.get(
        "group",
        ""
    ).strip()

    today = datetime.now().date()

    # --------------------------------
    # START QUERY
    # --------------------------------

    query = Attendance.query

    # --------------------------------
    # STUDENT FILTER
    # --------------------------------

    if student_id:

        query = query.filter(
            Attendance.student_id == student_id
        )

    # --------------------------------
    # BRANCH / GROUP FILTER
    # --------------------------------

    if selected_group:

        query = query.filter(
            Attendance.branch == selected_group
        )

    # --------------------------------
    # DATE FILTER
    # --------------------------------

    if report_type == "daily":

        query = query.filter(
            Attendance.date == today
        )

    elif report_type == "weekly":

        start_date = today - timedelta(days=7)

        query = query.filter(
            Attendance.date >= start_date
        )

    elif report_type == "monthly":

        start_date = today.replace(day=1)

        query = query.filter(
            Attendance.date >= start_date
        )

    # --------------------------------
    # GET RECORDS
    # --------------------------------

    records = query.order_by(
        Attendance.date.desc(),
        Attendance.time_in.desc()
    ).all()

    # --------------------------------
    # CREATE EXCEL
    # --------------------------------

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # --------------------------------
    # GROUP RECORDS BY BRANCH
    # --------------------------------

    groups = {}

    for record in records:

        branch = record.branch or "Unknown"

        if branch not in groups:
            groups[branch] = []

        groups[branch].append(record)

    # --------------------------------
    # NO DATA
    # --------------------------------

    if not groups:

        ws = wb.create_sheet("No Data")

        ws.append([
            "No attendance records found"
        ])

    # --------------------------------
    # CREATE BRANCH SHEETS
    # --------------------------------

    else:

        for branch, branch_records in groups.items():

            # Excel sheet name maximum = 31 characters
            sheet_name = branch[:31]

            ws = wb.create_sheet(sheet_name)

            # --------------------------------
            # CLEAN HEADER
            # --------------------------------

            ws.append([
                "Date",
                "Student ID",
                "Student Name",
                "Branch",
                "Topic",
                "System Number",
                "Remarks",
                "Time In"
            ])

            # --------------------------------
            # ADD ATTENDANCE DATA
            # --------------------------------

            for record in branch_records:

                ws.append([

                    str(record.date),

                    record.student_id,

                    record.student_name,

                    record.branch,

                    record.topic,

                    record.system_number,

                    record.remarks,

                    record.time_in.strftime("%I:%M %p")
                    if record.time_in
                    else ""

                ])

            # --------------------------------
            # FORMAT COLUMNS
            # --------------------------------

            column_widths = {

                "A": 14,
                "B": 15,
                "C": 25,
                "D": 12,
                "E": 30,
                "F": 18,
                "G": 35,
                "H": 15

            }

            for column, width in column_widths.items():

                ws.column_dimensions[column].width = width

            # --------------------------------
            # FREEZE HEADER
            # --------------------------------

            ws.freeze_panes = "A2"

            # --------------------------------
            # AUTO FILTER
            # --------------------------------

            ws.auto_filter.ref = ws.dimensions

    # --------------------------------
    # SAVE EXCEL TO MEMORY
    # --------------------------------

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    # --------------------------------
    # FILE NAME
    # --------------------------------

    if selected_group:

        filename = (
            f"{selected_group}_"
            f"{report_type.capitalize()}_"
            f"Attendance.xlsx"
        )

    else:

        filename = (
            f"All_Groups_"
            f"{report_type.capitalize()}_"
            f"Attendance.xlsx"
        )

    # --------------------------------
    # DOWNLOAD
    # --------------------------------

    return send_file(

        output,

        as_attachment=True,

        download_name=filename,

        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )

    )